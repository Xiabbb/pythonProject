import io

import requests
from gooey import Gooey, GooeyParser
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


@Gooey(
    richtext_controls=True,  # 打开终端对颜色支持
    language='chinese',
    header_show_title=False,
    program_name="单元格URL转图片工具v1.0",  # 程序名称
    encoding="utf-8",  # 设置编码格式，打包的时候遇到问题
    progress_regex=r"^progress: (\d+)%$",  # 正则，用于模式化运行时进度信息
    default_size=(650, 555),
    progress_expr="current / total * 100",
    # 再次执行，清除之前执行的日志
    clear_before_run=True,
    timing_options={'show_time_remaining': True, 'hide_time_remaining_on_complete': False}
)
def main():
    desc = f'工具说明:\n     本工具仅支持Excel表格内url转图片'
    file_help_msg = "help..."
    my_cool_parser = GooeyParser(description=desc)
    my_cool_parser.add_argument('filename', help='请选择要转换的文件(xlsx/XLS/CSV)', widget="FileChooser")
    my_cool_parser.add_argument('path', help='请选择保存目录', widget="DirChooser")
    # 获取参数
    args = my_cool_parser.parse_args()

    print(args, flush=True)  # 坑点：flush=True在打包的时候会用到
    return args


if __name__ == "__main__":
    args = main()
    filename = args.filename
    path = args.path
    wb = load_workbook(filename)
    ws = wb.active
    o = 0
    for row in ws.iter_rows():
        for cell in row:
            url = cell.value  # 获取单元格的值
            l = cell.coordinate  # 获取单元格坐标，例如 A1
            if 'jpg' in str(url) or 'png' in str(url) or 'jpeg' in str(url):
                o += 1
                ws.column_dimensions[cell.column_letter].width = 25.5  # 设置列宽
                ws.row_dimensions[cell.row].height = 115  # 设置行高
                res = requests.get(url)  # 下载图片
                img = Image(io.BytesIO(res.content))  # 获取图片内存地址
                img.width = 200  # 设置图片宽度
                img.height = 150  # 设 置图片高度

                ws.add_image(img, l)  # 将图片插入url所在的单元格
                print(f'正在下载并插入第 {o} 张图片', flush=True)
    wb.save(f'{path}/结果文件.xlsx')
    print('已完成')
